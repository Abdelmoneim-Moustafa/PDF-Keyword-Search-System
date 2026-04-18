
"""
PDF Keyword Search — Streamlit Cloud Edition

- No external storage service is required.
- Job metadata is persisted in a local SQLite database.
- Generated XLSX/CSV files are written to a local runtime folder so they can be downloaded.
- The UI is split into smart tabs with upload preview, live progress, and result summaries.
"""

from __future__ import annotations

import asyncio
import io
import json
import os
import sqlite3
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

import aiohttp
import pandas as pd
import streamlit as st
import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ──────────────────────────────────────────────────────────────────────────────
# Paths / Config
# ──────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
RUNTIME_DIR = BASE_DIR / "runtime"
RESULTS_DIR = RUNTIME_DIR / "results"
RUNTIME_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

DB_PATH = RUNTIME_DIR / "jobs.db"

CONCURRENCY = 150
MAX_BYTES = 524_288          # 512 KB range cap per PDF
TIMEOUT = 20
RETRIES = 2
LIMIT = 50_000
MAX_FEATURES = 3

UA = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0",
    "Accept": "application/pdf,*/*",
}

# ──────────────────────────────────────────────────────────────────────────────
# DB helpers
# ──────────────────────────────────────────────────────────────────────────────
def _db() -> sqlite3.Connection:
    return sqlite3.connect(DB_PATH, check_same_thread=False, timeout=15)


def init_db() -> None:
    with _db() as c:
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS jobs(
                id TEXT PRIMARY KEY,
                name TEXT,
                filename TEXT,
                started TEXT,
                status TEXT DEFAULT 'Pending',
                total INT DEFAULT 0,
                done INT DEFAULT 0,
                found INT DEFAULT 0,
                failed INT DEFAULT 0,
                xlsx_bytes BLOB,
                csv_bytes BLOB,
                xlsx_path TEXT,
                csv_path TEXT,
                err TEXT,
                progress TEXT DEFAULT '{}'
            )
            """
        )

        # Small migration helper for older databases.
        cols = {row[1] for row in c.execute("PRAGMA table_info(jobs)").fetchall()}
        for col, ddl in (
            ("xlsx_path", "ALTER TABLE jobs ADD COLUMN xlsx_path TEXT"),
            ("csv_path", "ALTER TABLE jobs ADD COLUMN csv_path TEXT"),
        ):
            if col not in cols:
                c.execute(ddl)


def save_job(j: dict) -> None:
    with _db() as c:
        c.execute(
            """
            INSERT OR REPLACE INTO jobs
            (id, name, filename, started, status, total, done, found, failed, err, progress)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                j["id"],
                j["name"],
                j["filename"],
                j["started"],
                j["status"],
                j["total"],
                0,
                0,
                0,
                "",
                "{}",
            ),
        )


def upd(jid: str, **kw) -> None:
    blobs = {k: v for k, v in kw.items() if k in ("xlsx_bytes", "csv_bytes")}
    plain = {k: v for k, v in kw.items() if k not in ("xlsx_bytes", "csv_bytes")}

    with _db() as c:
        if plain:
            c.execute(
                f"UPDATE jobs SET {','.join(k + '=?' for k in plain)} WHERE id=?",
                [*plain.values(), jid],
            )
        if blobs:
            for k, v in blobs.items():
                c.execute(f"UPDATE jobs SET {k}=? WHERE id=?", (v, jid))


def get_job(jid: str) -> dict | None:
    with _db() as c:
        cur = c.execute("SELECT * FROM jobs WHERE id=?", (jid,))
        row = cur.fetchone()
        if not row:
            return None
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


def all_jobs() -> pd.DataFrame:
    with _db() as c:
        return pd.read_sql(
            """
            SELECT id, name, filename, started, status, total, done, found, failed, err,
                   xlsx_path, csv_path
            FROM jobs
            ORDER BY started DESC
            """,
            c,
        )


def cleanup() -> None:
    cut = (datetime.now() - timedelta(days=7)).isoformat()
    with _db() as c:
        # delete old files first
        rows = c.execute(
            "SELECT xlsx_path, csv_path FROM jobs WHERE started<?", (cut,)
        ).fetchall()
        for xlsx_path, csv_path in rows:
            for p in (xlsx_path, csv_path):
                try:
                    if p and Path(p).exists():
                        Path(p).unlink()
                except Exception:
                    pass
        c.execute("DELETE FROM jobs WHERE started<?", (cut,))


# ──────────────────────────────────────────────────────────────────────────────
# Validation / Template
# ──────────────────────────────────────────────────────────────────────────────
def validate(df: pd.DataFrame) -> Tuple[List[str], List[str]]:
    errs, warns = [], []
    miss = [col for col in ("URL", "Keyword") if col not in df.columns]
    if miss:
        errs.append(f"Missing columns: {miss}")
        return errs, warns

    if df.empty:
        errs.append("File has no data rows.")
    if len(df) > LIMIT:
        errs.append(f"Exceeds {LIMIT:,} row limit ({len(df):,} rows)")

    nu = int(df["URL"].isna().sum())
    if nu:
        warns.append(f"{nu} empty URL rows will be skipped.")

    too_many = int(
        sum(
            1
            for x in df["Keyword"].fillna("").astype(str)
            if len([k for k in x.split("|") if k.strip()]) > MAX_FEATURES
        )
    )
    if too_many:
        warns.append(f"{too_many} row(s) contain more than {MAX_FEATURES} keywords; only the first {MAX_FEATURES} will be used.")
    return errs, warns


@st.cache_data
def make_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    hf = PatternFill("solid", fgColor="1F4E79")
    hfon = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    th = Side(style="thin", color="AAAAAA")
    bdr = Border(left=th, right=th, top=th, bottom=th)

    for i, h in enumerate(["URL", "Keyword"], 1):
        c = ws.cell(1, i, h)
        c.fill = hf
        c.font = hfon
        c.border = bdr
        c.alignment = Alignment(horizontal="center")

    dfon = Font(name="Arial", size=10)
    for i, (u, k) in enumerate(
        [
            ("https://example.com/doc1.pdf", "39131706"),
            ("https://example.com/doc2.pdf", "EAN13 8013975216323|HS Code 853"),
            ("https://example.com/doc3.pdf", "keyword1|keyword2|keyword3"),
        ],
        2,
    ):
        for j, v in enumerate([u, k], 1):
            c = ws.cell(i, j, v)
            c.font = dfon
            c.border = bdr

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 40
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Search engine
# ──────────────────────────────────────────────────────────────────────────────
def raw_search(data: bytes, keywords: list) -> dict:
    text = data.decode("latin-1", errors="replace")
    tlow = text.lower()
    out = {}
    for kw in keywords:
        idx = tlow.find(kw.lower())
        if idx >= 0:
            ctx = text[max(0, idx - 15): idx + len(kw) + 30].replace("\n", " ").strip()
            out[kw] = ("Found", ctx[:80])
        else:
            out[kw] = ("Not Found", None)
    return out


def pdf_search(data: bytes, keywords: list) -> dict:
    try:
        doc = fitz.open(stream=data, filetype="pdf")
        text = "".join(p.get_text() for p in doc)
        doc.close()
    except Exception:
        return {kw: ("Error", None) for kw in keywords}

    tlow = text.lower()
    out = {}
    for kw in keywords:
        idx = tlow.find(kw.lower())
        if idx >= 0:
            ctx = text[max(0, idx - 15): idx + len(kw) + 30].replace("\n", " ").strip()
            out[kw] = ("Found", ctx[:80])
        else:
            out[kw] = ("Not Found", None)
    return out


async def search_one(session, sem, url: str, keywords: list, pool) -> tuple:
    empty = {kw: ("Not Found", None) for kw in keywords}
    url = (url or "").strip()
    if not url or not url.startswith("http"):
        return empty, "Invalid URL"

    loop = asyncio.get_event_loop()

    for attempt in range(RETRIES + 1):
        try:
            async with sem:
                hdrs = {**UA, "Range": f"bytes=0-{MAX_BYTES - 1}"}
                tmo = aiohttp.ClientTimeout(total=TIMEOUT)
                async with session.get(
                    url, headers=hdrs, timeout=tmo, allow_redirects=True, ssl=False
                ) as r:
                    if r.status not in (200, 206):
                        if attempt < RETRIES:
                            await asyncio.sleep(1.5**attempt)
                            continue
                        return empty, f"HTTP {r.status}"
                    data = await r.read()

            if not data:
                return empty, "Empty"

            raw = await loop.run_in_executor(pool, raw_search, data, keywords)
            if all(v[0] == "Found" for v in raw.values()):
                return raw, "Done"

            if (b"FlateDecode" in data or b"flatedecode" in data) and any(v[0] != "Found" for v in raw.values()):
                full = await loop.run_in_executor(pool, pdf_search, data, keywords)
                merged = {kw: raw[kw] if raw[kw][0] == "Found" else full[kw] for kw in keywords}
                return merged, "Done"

            return raw, "Done"

        except asyncio.TimeoutError:
            if attempt < RETRIES:
                await asyncio.sleep(1)
                continue
            return empty, "Timeout"
        except Exception as e:
            if attempt < RETRIES:
                await asyncio.sleep(1)
                continue
            return empty, str(e)[:40]

    return empty, "Failed"


async def _run(jid: str, df: pd.DataFrame):
    job = get_job(jid)
    prog = json.loads(job.get("progress", "{}") or "{}")
    done_set = {int(k) for k in prog}
    total = len(df)

    upd(
        jid,
        status="Running",
        total=total,
        done=len(done_set),
        found=sum(1 for v in prog.values() if v.get("us") == "Done"),
        failed=len(done_set) - sum(1 for v in prog.values() if v.get("us") == "Done"),
    )

    sem = asyncio.Semaphore(CONCURRENCY)
    conn = aiohttp.TCPConnector(
        limit=CONCURRENCY, limit_per_host=30, ttl_dns_cache=300, enable_cleanup_closed=True
    )
    pool = ThreadPoolExecutor(max_workers=8)
    batch_size = 50

    async with aiohttp.ClientSession(connector=conn) as session:
        pending = []
        for idx, row in df.iterrows():
            if idx in done_set:
                continue
            url = str(row.get("URL", "")).strip()
            kws = [k.strip() for k in str(row.get("Keyword", "")).split("|") if k.strip()][:MAX_FEATURES]
            pending.append((idx, url, kws))

        for b in range(0, len(pending), batch_size):
            batch = pending[b:b + batch_size]
            res = await asyncio.gather(
                *[search_one(session, sem, url, kws, pool) for (_, url, kws) in batch],
                return_exceptions=True,
            )

            for (idx, url, kws), r in zip(batch, res):
                if isinstance(r, Exception):
                    kw_res, us = {kw: ("Not Found", None) for kw in kws}, str(r)[:40]
                else:
                    kw_res, us = r

                prog[str(idx)] = {
                    "us": us,
                    "kw": {k: {"s": v[0], "ctx": v[1]} for k, v in kw_res.items()},
                }

            n_done = len(prog)
            n_found = sum(1 for v in prog.values() if v.get("us") == "Done")
            upd(
                jid,
                total=total,
                done=n_done,
                found=n_found,
                failed=n_done - n_found,
                progress=json.dumps(prog),
            )

    pool.shutdown(wait=False)


def build_output(df: pd.DataFrame, prog: dict) -> pd.DataFrame:
    rows = []
    for idx, row in df.iterrows():
        kw_str = str(row.get("Keyword", "")).strip()
        kws = [k.strip() for k in kw_str.split("|") if k.strip()][:MAX_FEATURES]
        p = prog.get(str(idx), {})
        us = p.get("us", "Pending")
        kwr = p.get("kw", {})

        for kw in (kws or [kw_str]):
            r = kwr.get(kw, {"s": "Not Found", "ctx": None})
            rows.append(
                {
                    "URL": str(row.get("URL", "")),
                    "Keyword": kw_str,
                    "Extraction Option": None,
                    "URL_Status": 3 if us == "Done" else 0,
                    "URL_Search_Status": us,
                    "Keyword_Status": 3.0 if r["s"] == "Found" else 0.0,
                    "feature_name": kw,
                    "feature_value": r.get("ctx"),
                    "Keyword_Search_Status": r["s"],
                }
            )
    return pd.DataFrame(rows)


def to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    hf = PatternFill("solid", fgColor="1F4E79")
    hfon = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    th = Side(style="thin", color="CCCCCC")
    bdr = Border(left=th, right=th, top=th, bottom=th)
    gf = PatternFill("solid", fgColor="C6EFCE")
    rf = PatternFill("solid", fgColor="FFCCCC")

    for i, h in enumerate(df.columns, 1):
        c = ws.cell(1, i, h)
        c.fill = hf
        c.font = hfon
        c.border = bdr
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font = Font(name="Arial", size=9)
            c.border = bdr
            if ci == len(df.columns):
                if str(val) == "Found":
                    c.fill = gf
                    c.font = Font(name="Arial", size=9, bold=True, color="276221")
                elif str(val) == "Not Found":
                    c.fill = rf
                    c.font = Font(name="Arial", size=9, color="9C0006")

    widths = [55, 30, 18, 12, 18, 15, 30, 35, 22]
    for i, w in enumerate(widths[:len(df.columns)], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def _safe_filename(name: str) -> str:
    keep = []
    for ch in name:
        if ch.isalnum() or ch in ("-", "_", ".", " "):
            keep.append(ch)
        else:
            keep.append("_")
    return "".join(keep).strip().replace(" ", "_")


def run_job(jid: str, df: pd.DataFrame):
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(_run(jid, df))
        loop.close()

        job = get_job(jid)
        prog = json.loads(job.get("progress", "{}") or "{}")
        out = build_output(df, prog)

        # Write real files to the runtime folder for download.
        started = str(job.get("started", datetime.now().isoformat()))[:19].replace(":", "-")
        base = _safe_filename(f"{started}_{job.get('filename', 'results')}")
        job_dir = RESULTS_DIR / jid
        job_dir.mkdir(parents=True, exist_ok=True)

        xlsx_path = job_dir / f"{base}.xlsx"
        csv_path = job_dir / f"{base}.csv"

        xlsx = to_xlsx_bytes(out)
        csv_b = to_csv_bytes(out)

        xlsx_path.write_bytes(xlsx)
        csv_path.write_bytes(csv_b)

        upd(
            jid,
            status="Done",
            xlsx_bytes=xlsx,
            csv_bytes=csv_b,
            xlsx_path=str(xlsx_path),
            csv_path=str(csv_path),
        )

    except Exception as e:
        upd(jid, status="Failed", err=str(e)[:500])


# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="PDF Search", page_icon="🔍", layout="wide")

st.markdown(
    """
<style>
[data-testid="stAppViewContainer"]{background:#f5f7fa}
.block-container{padding:1.2rem 1.7rem;max-width:1400px}
.card{background:white;border-radius:16px;padding:18px 20px;
      box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:12px}
.big{font-size:2.2rem;font-weight:800;line-height:1}
.lbl{font-size:.78rem;color:#888;margin-top:4px}
.tag{display:inline-block;padding:3px 12px;border-radius:20px;
     font-size:.8rem;font-weight:700;letter-spacing:.3px}
.done   {background:#dcfce7;color:#166534}
.running{background:#fef9c3;color:#854d0e}
.failed {background:#fee2e2;color:#991b1b}
.pending{background:#f1f5f9;color:#475569}
.row-card{background:white;border-radius:14px;padding:14px 18px;
          box-shadow:0 1px 5px rgba(0,0,0,.06);margin-bottom:10px}
.small-note{color:#64748b;font-size:.92rem}
</style>
""",
    unsafe_allow_html=True,
)

init_db()
cleanup()

st.markdown(
    """
<div style='background:linear-gradient(135deg,#1e3a5f,#2563eb);
     padding:20px 28px;border-radius:16px;color:white;margin-bottom:18px'>
  <div style='font-size:1.6rem;font-weight:800'>🔍 PDF Keyword Search System</div>
  <div style='opacity:.85;margin-top:4px;font-size:.9rem'>
    Smart upload preview · Live progress · Download XLSX/CSV · No external storage service needed
  </div>
</div>
""",
    unsafe_allow_html=True,
)

tab_overview, tab_run, tab_jobs, tab_help = st.tabs(
    ["Overview", "Run Search", "Jobs", "Help"]
)

jobs = all_jobs()
running_exists = not jobs.empty and jobs["status"].isin(["Running", "Pending"]).any()

# ── Overview ──────────────────────────────────────────────────────────────────
with tab_overview:
    if jobs.empty:
        st.info("No jobs yet. Upload a file in the Run Search tab.")
    else:
        m1, m2, m3, m4, m5 = st.columns(5)
        metrics = [
            (len(jobs), "Total jobs", "#1e3a5f"),
            ((jobs.status == "Running").sum(), "Running", "#854d0e"),
            ((jobs.status == "Done").sum(), "Done", "#166534"),
            ((jobs.status == "Failed").sum(), "Failed", "#991b1b"),
            ((jobs.status == "Pending").sum(), "Pending", "#475569"),
        ]
        for col, (v, l, clr) in zip([m1, m2, m3, m4, m5], metrics):
            col.markdown(
                f"<div class='card' style='text-align:center;padding:14px'>"
                f"<div class='big' style='color:{clr}'>{v}</div>"
                f"<div class='lbl'>{l}</div></div>",
                unsafe_allow_html=True,
            )

        chart_df = pd.DataFrame(
            {
                "Status": ["Done", "Running", "Failed", "Pending"],
                "Count": [
                    int((jobs.status == "Done").sum()),
                    int((jobs.status == "Running").sum()),
                    int((jobs.status == "Failed").sum()),
                    int((jobs.status == "Pending").sum()),
                ],
            }
        ).set_index("Status")
        st.subheader("Job snapshot")
        st.bar_chart(chart_df)

        st.subheader("Latest jobs")
        st.dataframe(
            jobs[["name", "status", "total", "done", "found", "failed", "started"]],
            use_container_width=True,
            hide_index=True,
        )

# ── Run Search ────────────────────────────────────────────────────────────────
with tab_run:
    left, right = st.columns([1.1, 1])
    with left:
        st.download_button(
            "⬇ Download template",
            make_template(),
            file_name="search_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with right:
        if st.button("🔄 Refresh page", use_container_width=True):
            st.rerun()

    uf = st.file_uploader("Upload Excel file with URL and Keyword columns", type=["xlsx"])

    if uf:
        try:
            df = pd.read_excel(uf, dtype=str)
            errs, warns = validate(df)

            for w in warns:
                st.warning(w)
            for e in errs:
                st.error(f"❌ {e}")

            if not errs:
                df = df.fillna("")
                valid_urls = int(df["URL"].astype(str).str.startswith("http").sum())
                kw_counts = df["Keyword"].astype(str).apply(
                    lambda x: len([k for k in x.split("|") if k.strip()])
                )

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Rows", f"{len(df):,}")
                c2.metric("Valid URLs", f"{valid_urls:,}")
                c3.metric("Average keywords", f"{kw_counts.mean():.1f}")
                c4.metric("Max keywords", f"{int(kw_counts.max()):,}")

                st.caption("Input preview")
                st.dataframe(df.head(10), use_container_width=True)

                if st.checkbox("Show keyword count visualization", value=True):
                    st.bar_chart(kw_counts.value_counts().sort_index())

                est = max(1, int(len(df) / max(1, CONCURRENCY) * 7 / 60))
                st.info(f"Estimated run time: about {est} minute(s) for this upload.")

                if st.button("🚀 Start search", type="primary", use_container_width=True):
                    jid = str(uuid.uuid4())[:8]
                    ts = datetime.now().strftime("%Y%m%d%H%M%S%f")[:16]
                    name = f"{ts}_{os.path.splitext(uf.name)[0]}"
                    save_job(
                        {
                            "id": jid,
                            "name": name,
                            "filename": uf.name,
                            "started": datetime.now().isoformat(),
                            "status": "Pending",
                            "total": len(df),
                        }
                    )
                    threading.Thread(target=run_job, args=(jid, df.copy()), daemon=True).start()
                    st.success("Job started.")
                    st.rerun()

        except Exception as e:
            st.error(f"❌ {e}")

# ── Jobs ──────────────────────────────────────────────────────────────────────
with tab_jobs:
    jobs = all_jobs()
    if jobs.empty:
        st.info("No jobs available yet.")
    else:
        for _, job in jobs.iterrows():
            status = job["status"]
            tag_class = {"Done": "done", "Running": "running", "Failed": "failed"}.get(status, "pending")
            tot = int(job["total"] or 0)
            done_n = int(job["done"] or 0)
            found = int(job["found"] or 0)
            fail = int(job["failed"] or 0)

            with st.container():
                st.markdown('<div class="row-card">', unsafe_allow_html=True)
                a, b, c, d, e, f = st.columns([3.5, 1, 1, 1, 1, 2])

                nm = str(job["name"])
                a.markdown(f"**{nm[:58]}{'…' if len(nm) > 58 else ''}**")
                b.markdown(f"<span class='tag {tag_class}'>{status}</span>", unsafe_allow_html=True)
                c.metric("Rows", f"{tot:,}")
                d.metric("Found", f"{found:,}")
                e.metric("Failed", f"{fail:,}")

                if status == "Done":
                    jid = str(job["id"])
                    xlsx_path = job.get("xlsx_path")
                    csv_path = job.get("csv_path")
                    ts_part = str(job["started"])[:10].replace("-", "")
                    fname = f"{ts_part}_{job['filename']}"

                    xlsx_b = None
                    csv_b = None
                    if xlsx_path and Path(xlsx_path).exists():
                        xlsx_b = Path(xlsx_path).read_bytes()
                    elif get_job(jid).get("xlsx_bytes") is not None:
                        xlsx_b = get_job(jid).get("xlsx_bytes")

                    if csv_path and Path(csv_path).exists():
                        csv_b = Path(csv_path).read_bytes()
                    elif get_job(jid).get("csv_bytes") is not None:
                        csv_b = get_job(jid).get("csv_bytes")

                    f1, f2 = f.columns(2)
                    if xlsx_b:
                        f1.download_button(
                            "⬇ XLSX",
                            xlsx_b,
                            file_name=_safe_filename(fname.replace(".xlsx", "").replace(".csv", "")) + ".xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"x{jid}",
                        )
                    if csv_b:
                        f2.download_button(
                            "⬇ CSV",
                            csv_b,
                            file_name=_safe_filename(fname.replace(".xlsx", "").replace(".csv", "")) + ".csv",
                            mime="text/csv",
                            key=f"c{jid}",
                        )

                    if xlsx_path:
                        st.caption(f"Saved file: `{xlsx_path}`")
                        if st.checkbox(f"Show output preview for {jid}", key=f"prev_{jid}"):
                            try:
                                preview_df = pd.read_excel(io.BytesIO(xlsx_b), dtype=str)
                                st.dataframe(preview_df.head(10), use_container_width=True)
                            except Exception:
                                st.warning("Preview unavailable.")
                elif status == "Failed":
                    st.error(str(job.get("err", "Error"))[:120])
                elif status == "Running":
                    f.markdown(f"⏳ {done_n:,} / {tot:,}")
                    if tot > 0:
                        pct = done_n / tot
                        st.progress(pct, text=f"Processing {done_n:,} / {tot:,} ({pct*100:.1f}%)")
                else:
                    f.markdown("⏳ Queued")

                st.markdown("</div>", unsafe_allow_html=True)

    if running_exists:
        time.sleep(2.5)
        st.rerun()

# ── Help ──────────────────────────────────────────────────────────────────────
with tab_help:
    st.markdown("### How to use")
    st.write(
        """
        1. Download the template.
        2. Fill the Excel file with `URL` and `Keyword`.
        3. Use `|` to separate multiple keywords in one row.
        4. Upload the file and start the search.
        5. Download the generated XLSX or CSV when the job is done.
        """
    )

    st.markdown("### Output")
    st.write(
        """
        The app creates real result files inside the runtime folder and also exposes them as downloads.
        No external cloud storage is required.
        """
    )

    st.markdown("### Notes")
    st.write(
        f"""
        - Max rows: {LIMIT:,}
        - Max keywords per row: {MAX_FEATURES}
        - PDF byte range cap: {MAX_BYTES:,} bytes
        - Timeout: {TIMEOUT} seconds
        """
    )
